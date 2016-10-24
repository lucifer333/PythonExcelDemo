#openpyxl不支持xls,只支持修改xlsx
from openpyxl import load_workbook

#如果F:\\demo1.xlsx不存在会报错
excel=load_workbook("F:\\demo1.xlsx")
sheet=excel.active

sheet.cell(row=50,column=5).value="你是傻逼吗"

sheet.cell(row=60,column=10,value="你才是傻逼")

excel.save("F:\\demo1.xlsx")

